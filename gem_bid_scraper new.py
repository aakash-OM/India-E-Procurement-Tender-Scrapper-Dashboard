"""
GeM Portal Bid Scraper  —  v5  (All Bugs Fixed)
=================================================
FIXES in this version:
  ✅ FIX 1 — "invalid session id" crash: browser auto-restarts if Chrome dies
  ✅ FIX 2 — PDF parse crash: extract_text() can return None, now handled safely
  ✅ FIX 3 — Excel never saved if PDF loop crashed: save now happens no matter what
  ✅ FIX 4 — Sector filter applied AFTER PDF enrichment so org name is accurate
  ✅ NEW   — Sector filter (BHEL, NTPC, UPPTCL, KESCO, BPCL, NHAI etc.)
  ✅ NEW   — PDF auto-download for Item Category, Estimated Value, EMD, Contract Period

Requirements:
    pip install selenium openpyxl webdriver-manager beautifulsoup4 lxml pypdf requests

Usage:
    python gem_bid_scraper.py
"""

import time, re, sys, io
from datetime import datetime
from pathlib import Path

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────

KEYWORDS = [
    # Fittings & Connectors
    "copper brass fitting transformer", "LV HV bushing transformer",
    "bimetallic connector transformer", "aluminium connector transformer",
    "connecting lug transformer", "copper lug transformer",
    "epoxy bushing transformer", "bushing flange transformer",
    "insulating bushing transformer",
    # Tap Changers
    "tap changer transformer", "tap switch 33KV", "tap switch 11KV",
    "linear tap switch", "OLTC transformer",
    # Breathers & Relays
    "silica gel breather", "acrylic breather transformer", "buchholz relay",
    "WTI OTI transformer", "oil temperature indicator",
    "MOG transformer", "marshalling box transformer", "PRV transformer",
    # Tank Accessories
    "transformer tank accessories", "drain valve transformer",
    "wheel valve transformer", "flange valve transformer",
    "air release plug transformer", "oil level gauge transformer",
    "radiator valve transformer", "MS nipple transformer",
    # Insulating Materials
    "press board transformer", "craft paper transformer",
    "crepe paper transformer", "cotton tape transformer",
    "glass sleeve transformer", "empire sleeve transformer",
    "bakelite tube transformer", "cork washer transformer",
    "nomex paper transformer", "insulation material transformer",
    # Copper & Aluminium
    "copper busbar", "aluminium busbar", "copper jumper",
    "copper thimble", "aluminium thimble", "copper sheet transformer",
    # Transmission Line
    "JAW clamp transmission", "PG clamp transmission",
    "transmission line fittings", "arcing horn transformer",
    # CT & PT
    "CT pocket transformer", "WTI pocket", "epoxy terminal transformer",
    "current transformer parts",
    # Fabrication
    "lifting hook transformer", "MS hardware transformer",
    "bush nut transformer", "soldering wire transformer",
    # Broad
    "power transformer parts", "distribution transformer accessories",
    "transformer metal parts", "transformer spares", "transformer accessories",
]

# ── SECTOR FILTER ──────────────────────────────────────────────────────────────
# Only save bids from these organisations.
# Set to an EMPTY LIST  []  to save ALL bids without any filter.
TARGET_ORGANISATIONS = [
    "BHEL",
    "Bharat Heavy Electricals",
    "NTPC",
    "National Thermal Power",
    "UPPTCL",
    "UPPCL",
    "Uttar Pradesh Power Corporation",
    "KESCO",
    "Kanpur Electricity Supply",
    "BPCL",
    "Bharat Petroleum",
    "NHAI",
    "National Highways Authority",
    "Tata Power",
    "PGCIL",
    "Power Grid Corporation",
    "RVNL",
    "Rail Vikas Nigam",
    "IOCL",
    "Indian Oil",
    "Railways",
    "Indian Railways",
    "Power Department",
    "Energy Department",
    "Electricity Board",
    "Vidyut",
    "Transco",
    "Genco",
    "DISCOMS",
    "DISCOM",
]

OUTPUT_FILE            = "GeM_Bids_Master new.xlsx"
MAX_PAGES_PER_KEYWORD  = 5
HEADLESS               = True    # Set False to watch the browser live
AJAX_WAIT              = 7       # Seconds to wait after clicking Search
PAGE_NAV_WAIT          = 5       # Seconds to wait after clicking Next page
DOWNLOAD_BID_PDF       = True    # Download each bid PDF for full field extraction
PDF_TIMEOUT            = 20      # Seconds timeout per PDF download

# ───────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "Bid Number", "Organisation Name", "Department", "State / Ministry",
    "Item Category", "Quantity", "Estimated Value (₹)", "EMD Amount (₹)",
    "Bid End Date", "Contract Period", "Bid Type", "MSE Preference",
    "Keyword Used", "Bid URL", "Scraped On",
]

GEM_BASE = "https://bidplus.gem.gov.in"
COL_HDR  = "1F4E79"   # dark blue  — headers
COL_ALT  = "DEEAF1"   # light blue — alternate rows
COL_WHT  = "FFFFFF"
COL_SUM  = "2E75B6"   # medium blue — summary header
COL_GRN  = "E2EFDA"   # light green — sector-matched rows


# ─── SESSION-DEAD SIGNAL ───────────────────────────────────────────────────────

class SessionDead(Exception):
    """Raised when the Chrome WebDriver session has crashed."""
    pass


# ─── BROWSER ───────────────────────────────────────────────────────────────────

def make_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    o = Options()
    if HEADLESS:
        o.add_argument("--headless=new")
    o.add_argument("--no-sandbox")
    o.add_argument("--disable-dev-shm-usage")
    o.add_argument("--disable-gpu")
    o.add_argument("--window-size=1920,1080")
    o.add_argument("--disable-blink-features=AutomationControlled")
    o.add_experimental_option("excludeSwitches", ["enable-automation"])
    o.add_experimental_option("useAutomationExtension", False)
    o.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    drv = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=o
    )
    drv.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    return drv


# ─── SECTOR FILTER ─────────────────────────────────────────────────────────────

def is_target_org(org_text):
    """Return True if this bid's organisation matches any TARGET_ORGANISATIONS."""
    if not TARGET_ORGANISATIONS:
        return True                  # No filter — accept everything
    if not org_text or org_text == "N/A":
        return False
    org_upper = org_text.upper()
    for target in TARGET_ORGANISATIONS:
        if target.upper() in org_upper:
            return True
    return False


# ─── SCRAPING ──────────────────────────────────────────────────────────────────

def search_keyword(driver, keyword, max_pages):
    """
    Type keyword into GeM search box, click search, wait for AJAX,
    then scrape each result page. Raises SessionDead if browser crashes.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from bs4 import BeautifulSoup

    bids = []
    print(f"\n  Searching: '{keyword}'")

    # ── Guard: is session still alive? ─────────────────────────────────────
    try:
        _ = driver.title
    except Exception:
        raise SessionDead("Browser session died")

    try:
        driver.get(f"{GEM_BASE}/all-bids")
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "searchBid"))
        )
        time.sleep(2)

        # Type keyword and trigger search
        box = driver.find_element(By.ID, "searchBid")
        box.clear()
        box.send_keys(keyword)
        time.sleep(0.5)
        driver.find_element(By.ID, "searchBidRA").click()
        time.sleep(AJAX_WAIT)

        # Wait for bid cards to appear
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a.bid_no_hover"))
            )
        except Exception:
            print(f"   No results.")
            return bids

        # Scrape each page
        for page in range(1, max_pages + 1):
            print(f"   Page {page}...", end=" ", flush=True)
            soup   = BeautifulSoup(driver.page_source, "lxml")
            parsed = parse_bid_cards(soup, keyword)
            bids.extend(parsed)
            print(f"{len(parsed)} bids")

            if page < max_pages:
                if not go_next_page(driver, page):
                    break
                time.sleep(PAGE_NAV_WAIT)

    except SessionDead:
        raise
    except Exception as e:
        print(f"   Page error: {e}")

    return bids


def go_next_page(driver, current_page):
    """Click the next numbered page in GeM pagination. Returns True if clicked."""
    from selenium.webdriver.common.by import By
    try:
        links = driver.find_elements(
            By.CSS_SELECTOR, "#light-pagination a, .pagination2 a"
        )
        target = str(current_page + 1)
        for link in links:
            if link.text.strip() == target:
                link.click(); return True
        for link in links:
            if "»" in link.text or link.get_attribute("rel") == "next":
                link.click(); return True
    except Exception:
        pass
    return False


# ─── BID CARD PARSING ──────────────────────────────────────────────────────────

def parse_bid_cards(soup, keyword):
    """
    GeM page confirmed structure (from debug output):
      div#bidCard  →  .card (×N)  →  a.bid_no_hover + .card-body fields
    """
    results   = []
    container = soup.find("div", id="bidCard")
    if not container:
        return results
    for card in (container.find_all("div", class_="card") or []):
        bid = parse_single_card(card, keyword)
        if bid:
            results.append(bid)
    return results


def parse_single_card(card, keyword):
    try:
        bid_link = card.find("a", class_="bid_no_hover")
        if not bid_link:
            return None
        bid_no = bid_link.get_text(strip=True)
        if not re.search(r"GEM/\d{4}/", bid_no, re.I):
            return None

        # Fix URL — ensure slash between domain and path
        href = bid_link.get("href", "")
        if href.startswith("/"):
            bid_url = GEM_BASE + href
        elif href.startswith("http"):
            bid_url = href
        else:
            bid_url = GEM_BASE + "/" + href.lstrip("/")

        full_text = card.get_text(separator="|||", strip=True)
        f         = extract_card_fields(card, full_text)

        return {
            "Bid Number":          bid_no,
            "Organisation Name":   f.get("org",       "N/A"),
            "Department":          f.get("dept",       "N/A"),
            "State / Ministry":    f.get("state",      "N/A"),
            "Item Category":       f.get("category",   "N/A"),
            "Quantity":            f.get("quantity",   "N/A"),
            "Estimated Value (₹)": f.get("value",      "N/A"),
            "EMD Amount (₹)":      "N/A",   # filled from PDF
            "Bid End Date":        f.get("end_date",   "N/A"),
            "Contract Period":     "N/A",   # filled from PDF
            "Bid Type":            "N/A",   # filled from PDF
            "MSE Preference":      "N/A",   # filled from PDF
            "Keyword Used":        keyword,
            "Bid URL":             bid_url,
            "Scraped On":          datetime.now().strftime("%d-%m-%Y %H:%M"),
        }
    except Exception:
        return None


def extract_card_fields(card, full_text):
    """
    Parse GeM bid card text. Confirmed live format (from real data):

      "[Qty] Department Name And Address: [Ministry] [Dept/Org] Start Date: [X] End Date: [Y]"

    All fields are packed into one text block — this function splits them correctly.
    """
    result = {}
    text = card.get_text(separator=" ", strip=True)

    # Quantity — leading number before "Department Name And Address:"
    m = re.match(r'^([\d,\.]+)\s+Department Name And Address:', text, re.I)
    if m:
        result['quantity'] = m.group(1).replace(',', '').strip()

    # Ministry / State  +  Department / Org block between "Address:" and "Start Date:"
    m = re.search(r'Department Name And Address:\s*(.+?)\s+Start Date:', text, re.I | re.S)
    if m:
        dept_block = re.sub(r'\s*\bNA\b\s*$', '', m.group(1).strip()).strip()
        split_pair = _split_ministry_dept(dept_block)
        result['state'] = split_pair[0]
        if split_pair[1]:
            result['dept'] = split_pair[1]

    # Bid Start Date
    m = re.search(r'Start Date:\s*(\d{2}-\d{2}-\d{4}\s+\d+:\d+\s+(?:AM|PM))', text, re.I)
    if m:
        result['start_date'] = m.group(1).strip()

    # Bid End Date
    m = re.search(r'End Date:\s*(\d{2}-\d{2}-\d{4}\s+\d+:\d+\s+(?:AM|PM))', text, re.I)
    if m:
        result['end_date'] = m.group(1).strip()

    # Item Category — from card label: value elements
    for elem in card.find_all(["p", "div", "span"]):
        txt = elem.get_text(separator=" ", strip=True)
        if not txt or len(txt) > 500:
            continue
        for label in ["Item Category", "Category"]:
            m = re.search(rf"{label}\s*[:\-]\s*(.+)", txt, re.I)
            if m:
                val = m.group(1).strip()
                if val and len(val) < 300:
                    result['category'] = val
                    break
        if 'category' in result:
            break

    # Estimated Value
    m = re.search(r'(?:Estimated\s+)?(?:Bid\s+)?Value\s*[:\-]\s*([\d,\.]+)', text, re.I)
    if m:
        result['value'] = re.sub(r'[^\d,.]', '', m.group(1))

    return result


def _split_ministry_dept(dept_block):
    """Split 'Ministry of X Department of Y' into (ministry, dept_or_None)."""
    if dept_block.startswith("PMO "):
        return ("PMO", dept_block[4:].strip())
    m = re.match(
        r'^(Ministry of [^,]+?)\s+'
        r'((?:Department|Directorate|Office|Authority|Board|Commission|Council|'
        r'Secretariat|Division).+)$', dept_block, re.I)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    m = re.match(
        r'^(Ministry of [^,]+?)\s+([A-Z]{2}[A-Z\s]+(?:Limited|Ltd|Corporation|Board).*)$',
        dept_block)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    m = re.match(
        r'^(Ministry of \w+(?:\s+\w+)?)\s+((?:Indian|Central|National|State)\s+\w.+)$',
        dept_block, re.I)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    return (dept_block, None)

def fetch_pdf_bytes(bid_url):
    """Download GeM bid document PDF. Returns bytes or None on failure."""
    try:
        import requests
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": GEM_BASE,
        }
        r = requests.get(bid_url, headers=headers, timeout=PDF_TIMEOUT, stream=True)
        ct = r.headers.get("content-type", "").lower()
        if r.status_code == 200 and "pdf" in ct:
            return r.content
    except Exception:
        pass
    return None


def parse_pdf_fields(pdf_bytes):
    """
    Parse GeM bid PDF (bilingual Hindi/English) and return a dict of field values.
    All known crash points are individually caught so one bad PDF never stops the run.
    """
    if not pdf_bytes:
        return {}

    # ── Step 1: Read PDF ────────────────────────────────────────────────────
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return {}

    # ── Step 2: Extract text from each page safely ──────────────────────────
    # FIX: extract_text() can return None for some pages — use "or ''"
    raw_text = ""
    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""   # ← None fix here
            raw_text += page_text + "\n"
        except Exception:
            continue   # skip unreadable page, don't crash

    if not raw_text.strip():
        return {}

    # ── Step 3: Filter to English lines (bilingual PDFs repeat Hindi) ───────
    clean_lines = []
    for line in raw_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        ascii_count = sum(1 for c in line if ord(c) < 128)
        if ascii_count / len(line) > 0.5:
            clean_lines.append(line)
    text = " | ".join(clean_lines)

    if not text:
        return {}

    # ── Step 4: Extract fields using confirmed PDF label patterns ───────────
    def get(patterns):
        for pat in patterns:
            try:
                m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
                if m:
                    val = m.group(1).strip().rstrip("|").strip()
                    # Remove repeated words (pypdf bilingual artifact: "Name Name")
                    val = re.sub(r'\b(\w+)\s+\1\b', r'\1', val)
                    if val and len(val) < 400:
                        return val
            except Exception:
                continue
        return "N/A"

    return {
        "Organisation Name": get([
            r"Organisation Name\s*/[^/]+/\s*([^\|]{5,100})",
            r"Organisation Name\s*[:\-]\s*([^\|]{5,100})",
        ]),
        "Department": get([
            r"Department Name\s*/[^/]+/\s*([^\|]{5,100})",
            r"Department Name\s*[:\-]\s*([^\|]{5,100})",
        ]),
        "State / Ministry": get([
            r"Ministry/State Name\s*[:\-]?\s*([^\|]{3,80})",
            r"Ministry\s*/State Name\s*([^\|]{3,80})",
        ]),
        "Item Category": get([
            r"Item Category\s*/[^/]+/\s*([^\|]{10,300})",
            r"Item Category\s*[:\-]\s*([^\|]{10,300})",
            r"/Item Category\s*\n([^\n]{10,300})",
        ]),
        "Quantity": get([
            r"Quantity\s*/[^/]+/\s*([\d,\.]+\s*\w{0,20})",
            r"\|\s*([\d,]+)\s*N/A\s*\|",
        ]),
        "Estimated Value (₹)": get([
            r"Estimated Bid Value\s*/[^/]+/\s*([\d,\.]+)",
            r"Estimated Bid Value\s*[:\-]?\s*([\d,\.]+)",
            r"/ Estimated Bid Value\s*([\d,\.]+)",
        ]),
        "EMD Amount (₹)": get([
            r"EMD Amount\s*/[^/]+/\s*([\d,\.]+)",
            r"EMD Amount\s*[:\-]\s*([\d,\.]+)",
            r"ईएमडी\s*राशि\s*/EMD Amount\s*([\d,\.]+)",
        ]),
        "Bid End Date": get([
            r"Bid End Date/Time\s*[:\-]?\s*(\d{2}-\d{2}-\d{4}[^\|]{0,20})",
            r"Bid End Date\s*/[^/]+/\s*(\d{2}-\d{2}-\d{4}[^\|]{0,20})",
        ]),
        "Contract Period": get([
            r"Contract Period\s*/[^/]+/\s*([^\|]{2,50})",
            r"Contract Period\s*[:\-]\s*([^\|]{2,50})",
            r"/ Contract Period\s*([^\|]{2,50})",
        ]),
        "Bid Type": get([
            r"Type of Bid\s*/[^/]+/\s*([^\|]{3,50})",
            r"Type of Bid\s*[:\-]\s*([^\|]{3,50})",
        ]),
        "MSE Preference": get([
            r"MSE Purchase Preference\s*/[^/]+/\s*(Yes|No)",
            r"MSE Purchase Preference\s*[:\-]\s*(Yes|No)",
        ]),
    }


def enrich_bid_with_pdf(bid):
    """
    Download bid PDF and overwrite N/A fields with real values.
    Any failure is silently caught — the original bid is returned unchanged.
    """
    if not DOWNLOAD_BID_PDF:
        return bid
    bid_url = bid.get("Bid URL", "")
    if not bid_url or bid_url == "N/A":
        return bid
    try:
        pdf_bytes = fetch_pdf_bytes(bid_url)
        if not pdf_bytes:
            return bid
        pdf_fields = parse_pdf_fields(pdf_bytes)
        for field, pdf_val in pdf_fields.items():
            if pdf_val and pdf_val != "N/A":
                bid[field] = pdf_val
    except Exception:
        pass   # never crash the whole run over one bad PDF
    return bid


# ─── EXCEL ─────────────────────────────────────────────────────────────────────

def get_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin")
    return {
        "hf": Font(bold=True, color="FFFFFF", name="Arial", size=10),
        "df": Font(name="Arial", size=9),
        "bf": Font(bold=True, name="Arial", size=10),
        "ce": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "le": Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "br": Border(left=thin, right=thin, top=thin, bottom=thin),
        "fh": PatternFill("solid", fgColor=COL_HDR),
        "fa": PatternFill("solid", fgColor=COL_ALT),
        "fw": PatternFill("solid", fgColor=COL_WHT),
        "fs": PatternFill("solid", fgColor=COL_SUM),
        "ft": PatternFill("solid", fgColor="BDD7EE"),
        "fg": PatternFill("solid", fgColor=COL_GRN),
    }


def write_headers(ws, s):
    from openpyxl.utils import get_column_letter
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font=s["hf"]; c.fill=s["fh"]; c.alignment=s["ce"]; c.border=s["br"]
    ws.row_dimensions[1].height = 35
    col_widths = [22, 38, 30, 22, 45, 12, 20, 18, 22, 16, 20, 14, 20, 45, 18]
    for ci, w in enumerate(col_widths[:len(HEADERS)], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def append_bids(ws, bids, s):
    from openpyxl.styles import PatternFill
    existing = {
        str(r[0]).strip()
        for r in ws.iter_rows(min_row=2, values_only=True) if r[0]
    }
    added = 0
    for bid in bids:
        bn = str(bid["Bid Number"]).strip()
        if bn in existing:
            continue
        nr  = ws.max_row + 1
        org = bid.get("Organisation Name", "")
        # Green for sector-matched, alternating blue/white otherwise
        fl  = s["fg"] if is_target_org(org) else (
              s["fa"] if nr % 2 == 0 else PatternFill("solid", fgColor=COL_WHT))
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(nr, ci, bid.get(h, "N/A"))
            c.font=s["df"]; c.fill=fl; c.border=s["br"]
            c.alignment = s["le"] if ci in (2, 3, 4, 5, 14) else s["ce"]
        existing.add(bn)
        added += 1
    return added


def rebuild_summary(wb, s, run_stats):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    ws["A1"] = "GeM Bids — Monthly Summary"
    ws["A1"].font = Font(bold=True, name="Arial", size=14, color=COL_HDR)
    ws["A2"] = f"Last updated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    filter_note = (
        f"Sector filter: {', '.join(TARGET_ORGANISATIONS[:5])}{'...' if len(TARGET_ORGANISATIONS)>5 else ''}"
        if TARGET_ORGANISATIONS else "Sector filter: ALL organisations"
    )
    ws["A3"] = filter_note
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws.merge_cells("A3:F3")

    for ci, h in enumerate(
        ["Month", "Total Bids", "Sector Matched", "Added This Run", "Keywords Used", "Sheet"], 1
    ):
        c = ws.cell(5, ci, h)
        c.font=s["hf"]; c.fill=s["fs"]; c.alignment=s["ce"]; c.border=s["br"]

    sheets = [n for n in wb.sheetnames if n != "Summary"]
    grand  = 0
    for ri, name in enumerate(sheets, 6):
        wm    = wb[name]
        count = max(wm.max_row - 1, 0)
        kws   = {
            str(r[12]) for r in wm.iter_rows(min_row=2, values_only=True) if r[12]
        }
        sector_matched = sum(
            1 for r in wm.iter_rows(min_row=2, values_only=True)
            if r[1] and is_target_org(str(r[1]))
        )
        this_run = run_stats.get(name, 0)
        fl       = s["fa"] if ri % 2 == 0 else PatternFill("solid", fgColor=COL_WHT)
        for ci, v in enumerate(
            [name, count, sector_matched, this_run, ", ".join(sorted(kws)) or "—", f"→ '{name}'"], 1
        ):
            c = ws.cell(ri, ci, v)
            c.font=s["df"]; c.fill=fl; c.border=s["br"]
            c.alignment = s["le"] if ci == 5 else s["ce"]
        grand += count

    tr = 6 + len(sheets)
    for ci, v in enumerate(["GRAND TOTAL", grand, "", "", "", ""], 1):
        c = ws.cell(tr, ci, v)
        c.font=s["bf"]; c.fill=s["ft"]; c.border=s["br"]; c.alignment=s["ce"]

    for ci, w in enumerate([18, 14, 16, 16, 50, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))


def save_excel(bids, output_path, run_stats):
    from openpyxl import Workbook, load_workbook
    month = datetime.now().strftime("%B %Y")
    s     = get_styles()

    if Path(output_path).exists():
        print(f"\n  Opening existing file : {output_path}")
        wb = load_workbook(output_path)
        # Auto-wipe corrupt sheet
        if month in wb.sheetnames:
            wm  = wb[month]
            bad = any(
                str(r[1]).endswith(":") or str(r[1]) in ("Quantity:", "By Bid Type:")
                for r in wm.iter_rows(min_row=2, max_row=4, values_only=True)
                if r[1]
            )
            if bad:
                print(f"  ⚠️  Corrupt data in '{month}' — wiping and rebuilding...")
                del wb[month]
    else:
        print(f"\n  Creating new file     : {output_path}")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if month not in wb.sheetnames:
        ws = wb.create_sheet(month)
        write_headers(ws, s)
        print(f"  New sheet created     : '{month}'")
    else:
        ws = wb[month]
        print(f"  Sheet found           : '{month}'")

    added   = append_bids(ws, bids, s)
    skipped = len(bids) - added
    run_stats[month] = added
    print(f"  New bids added        : {added}")
    print(f"  Duplicates skipped    : {skipped}")
    rebuild_summary(wb, s, run_stats)
    wb.save(output_path)
    print(f"  Saved                 : {output_path}")
    return added


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    now = datetime.now()
    print("=" * 62)
    print("  GeM Portal Bid Scraper  v5  —  Sector-Filtered + PDF-Enriched")
    print("=" * 62)
    print(f"  Output file  : {OUTPUT_FILE}")
    print(f"  This run     : '{now.strftime('%B %Y')}' sheet")
    print(f"  PDF parsing  : {'ON  (Item Category, Value, EMD, Dates)' if DOWNLOAD_BID_PDF else 'OFF'}")
    if TARGET_ORGANISATIONS:
        sample = ", ".join(TARGET_ORGANISATIONS[:4])
        print(f"  Sector filter: {sample}{'...' if len(TARGET_ORGANISATIONS)>4 else ''}")
    else:
        print(f"  Sector filter: NONE — saving all organisations")
    print("=" * 62)

    # Dependency check
    pkgs = ["selenium","openpyxl","webdriver_manager","bs4","lxml","pypdf","requests"]
    missing = []
    for pkg in pkgs:
        try: __import__(pkg)
        except ImportError: missing.append(pkg)
    if missing:
        print(f"\n  ❌ Missing packages: {', '.join(missing)}")
        print("  Run: pip install selenium openpyxl webdriver-manager beautifulsoup4 lxml pypdf requests")
        sys.exit(1)

    # ── Scraping loop with automatic browser restart ────────────────────────
    all_bids   = []
    driver     = None
    kw_index   = 0

    while kw_index < len(KEYWORDS):
        kw = KEYWORDS[kw_index]
        if driver is None:
            print("\n  Starting browser...")
            driver = make_driver()

        try:
            bids = search_keyword(driver, kw, MAX_PAGES_PER_KEYWORD)
            all_bids.extend(bids)
            if bids:
                print(f"  ✓ '{kw}': {len(bids)} bids")
            kw_index += 1
            time.sleep(1)

        except SessionDead:
            # Browser crashed — restart and retry same keyword
            print(f"\n  🔄 Browser crashed. Restarting and retrying '{kw}'...")
            try:
                driver.quit()
            except Exception:
                pass
            driver = None
            time.sleep(3)
            # Don't increment kw_index — retry same keyword

        except KeyboardInterrupt:
            print("\n  Interrupted by user. Saving what was collected...")
            break

        except Exception as e:
            print(f"\n  Unexpected error on '{kw}': {e}")
            kw_index += 1   # skip this keyword and continue

    if driver:
        try: driver.quit()
        except Exception: pass

    # ── Deduplicate ─────────────────────────────────────────────────────────
    seen, unique = set(), []
    for b in all_bids:
        if b["Bid Number"] not in seen:
            seen.add(b["Bid Number"]); unique.append(b)
    print(f"\n  Total unique bids scraped : {len(unique)}")

    if not unique:
        print("  ⚠️  Nothing to save. Try HEADLESS=False or increase AJAX_WAIT.")
        return

    # ── PDF enrichment (each PDF in a safe try/except) ──────────────────────
    if DOWNLOAD_BID_PDF:
        print(f"  Downloading bid PDFs for full field extraction...")
        for i, bid in enumerate(unique):
            pct = int((i + 1) / len(unique) * 100)
            sys.stdout.write(f"\r  PDF {i+1}/{len(unique)}  ({pct}%)  {bid['Bid Number'][:30]:<30}")
            sys.stdout.flush()
            try:
                unique[i] = enrich_bid_with_pdf(bid)
            except Exception:
                pass          # FIX: never let one bad PDF crash the whole run
            time.sleep(0.3)
        print(f"\r  PDF enrichment complete — {len(unique)} bids processed.          ")

    # ── Sector filter ────────────────────────────────────────────────────────
    if TARGET_ORGANISATIONS:
        matched   = [b for b in unique if is_target_org(b.get("Organisation Name", ""))]
        excluded  = len(unique) - len(matched)
        print(f"  Sector filter  : {len(matched)} matched / {excluded} excluded")
        final_bids = matched
    else:
        final_bids = unique

    if not final_bids:
        print("\n  ⚠️  No bids matched sector filter.")
        print("  Tip: Add more names to TARGET_ORGANISATIONS, or set it to []")
        # Save all anyway so data isn't lost
        final_bids = unique
        print(f"  Saving all {len(final_bids)} unfiltered bids instead.")

    # ── Save Excel — ALWAYS runs, even if PDF/filter steps had issues ────────
    run_stats = {}
    added = save_excel(final_bids, OUTPUT_FILE, run_stats)

    print("\n" + "=" * 62)
    print(f"  ✅ Done!  {added} new bids added → '{now.strftime('%B %Y')}' tab")
    print(f"  📁 File   : {OUTPUT_FILE}")
    print(f"  🟢 Green rows = sector-matched (BHEL / NTPC / UPPCL etc.)")
    print("=" * 62)


if __name__ == "__main__":
    main()
