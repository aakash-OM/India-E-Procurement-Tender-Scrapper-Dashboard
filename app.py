"""
Gov Tender Dashboard – Flask backend
"""
import json
import os
import threading
import io
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file

from database import Database
from portals_config import ALL_PORTALS, PORTAL_BY_ID, REGIONS

app = Flask(__name__)
db = Database()

# Shared scraping state (single-worker model)
scrape_status = {
    "active":          False,
    "progress":        0,
    "pdf_progress":    0,
    "current_portal":  "",
    "current_keyword": "",
    "bids_found":      0,
    "job_id":          None,
    "log":             [],
}


# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── Keywords API ──────────────────────────────────────────────────────────────

@app.route("/api/keywords", methods=["GET"])
def get_keywords():
    return jsonify(db.get_keywords())


@app.route("/api/keywords", methods=["POST"])
def add_keyword():
    data = request.get_json(force=True)
    keyword  = (data.get("keyword") or "").strip()
    category = (data.get("category") or "General").strip()
    if not keyword:
        return jsonify({"error": "keyword required"}), 400
    kid = db.add_keyword(keyword, category)
    return jsonify({"id": kid, "keyword": keyword, "category": category, "is_active": 1}), 201


@app.route("/api/keywords/<int:kid>", methods=["DELETE"])
def delete_keyword(kid):
    db.delete_keyword(kid)
    return jsonify({"success": True})


@app.route("/api/keywords/<int:kid>/toggle", methods=["POST"])
def toggle_keyword(kid):
    db.toggle_keyword(kid)
    return jsonify({"success": True})


# ── Portals API ───────────────────────────────────────────────────────────────

@app.route("/api/portals", methods=["GET"])
def get_portals():
    return jsonify(ALL_PORTALS)


@app.route("/api/regions", methods=["GET"])
def get_regions():
    return jsonify(REGIONS)


# ── Bids API ──────────────────────────────────────────────────────────────────

@app.route("/api/bids", methods=["GET"])
def get_bids():
    result = db.get_bids(
        state     = request.args.get("state"),
        region    = request.args.get("region"),
        portal    = request.args.get("portal"),
        keyword   = request.args.get("keyword"),
        date_from = request.args.get("date_from"),
        date_to   = request.args.get("date_to"),
        min_value = float(request.args["min_value"]) if request.args.get("min_value") else None,
        max_value = float(request.args["max_value"]) if request.args.get("max_value") else None,
        limit     = int(request.args.get("limit", 100)),
        offset    = int(request.args.get("offset", 0)),
    )
    return jsonify(result)


@app.route("/api/bids/export", methods=["GET"])
def export_bids():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    result = db.get_bids(
        state   = request.args.get("state"),
        region  = request.args.get("region"),
        portal  = request.args.get("portal"),
        keyword = request.args.get("keyword"),
        limit   = 10000,
    )
    bids = result["bids"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenders"

    headers = [
        "Bid Number", "Portal", "Organisation", "Department",
        "State", "Region", "Item Category", "Quantity",
        "Est. Value (₹)", "Start Date", "End Date",
        "Keyword", "URL", "Scraped On"
    ]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill("solid", fgColor="DEEAF1")
    for i, bid in enumerate(bids):
        row = [
            bid.get("bid_number", ""), bid.get("portal_id", ""),
            bid.get("org_name", ""), bid.get("department", ""),
            bid.get("state", ""), bid.get("region", ""),
            bid.get("item_category", ""), bid.get("quantity", ""),
            bid.get("estimated_value"), bid.get("bid_start_date", ""),
            bid.get("bid_end_date", ""), bid.get("keyword_used", ""),
            bid.get("bid_url", ""), bid.get("scraped_at", ""),
        ]
        ws.append(row)
        if i % 2 == 1:
            for cell in ws[i + 2]:
                cell.fill = alt_fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"GovTenders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Stats API ─────────────────────────────────────────────────────────────────

@app.route("/api/stats", methods=["GET"])
def get_stats():
    return jsonify(db.get_stats())


# ── Search / Scraping API ─────────────────────────────────────────────────────

@app.route("/api/search", methods=["POST"])
def start_search():
    global scrape_status
    if scrape_status["active"]:
        return jsonify({"error": "Scraping already in progress"}), 409

    data        = request.get_json(force=True)
    portals     = data.get("portals", ["gem"])
    keyword_ids = data.get("keyword_ids", [])
    target_orgs = data.get("target_orgs", [])
    max_pages   = int(data.get("max_pages", 5))
    headless    = bool(data.get("headless", True))
    filter_states = data.get("states", [])

    # Resolve keywords from DB
    all_kws = db.get_keywords()
    if keyword_ids:
        keywords = [k["keyword"] for k in all_kws if k["id"] in keyword_ids and k["is_active"]]
    else:
        keywords = [k["keyword"] for k in all_kws if k["is_active"]]

    if not keywords:
        return jsonify({"error": "No active keywords"}), 400

    job_id = db.create_job(portals, keywords)
    scrape_status.update({
        "active": True, "progress": 0, "pdf_progress": 0,
        "bids_found": 0, "job_id": job_id, "log": [],
        "current_portal": "", "current_keyword": "",
    })

    def worker():
        global scrape_status
        total_saved = 0
        try:
            if "gem" in portals:
                from scrapers.gem_scraper import run_gem_scrape
                saved = run_gem_scrape(keywords, target_orgs, max_pages,
                                       scrape_status, db, headless=headless)
                total_saved += saved

            # Placeholder for future portal scrapers
            for portal_id in portals:
                if portal_id == "gem":
                    continue
                p = PORTAL_BY_ID.get(portal_id, {})
                scrape_status["log"].append(
                    f"[{p.get('name', portal_id)}] Support coming soon — portal not yet implemented."
                )

            db.update_job(job_id,
                          status="completed",
                          bids_found=total_saved,
                          completed_at=datetime.now().isoformat())
        except Exception as e:
            scrape_status["log"].append(f"[ERROR] {e}")
            db.update_job(job_id, status="error", error_msg=str(e),
                          completed_at=datetime.now().isoformat())
        finally:
            scrape_status["active"]   = False
            scrape_status["progress"] = 100

    t = threading.Thread(target=worker, daemon=True)
    t.start()
    return jsonify({"success": True, "job_id": job_id})


@app.route("/api/search/status", methods=["GET"])
def search_status():
    return jsonify(scrape_status)


@app.route("/api/search/stop", methods=["POST"])
def stop_search():
    scrape_status["active"] = False
    return jsonify({"success": True})


@app.route("/api/jobs", methods=["GET"])
def get_jobs():
    return jsonify(db.get_recent_jobs())


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = not os.environ.get("RAILWAY_ENVIRONMENT")  # debug off on Railway
    print(f"\n  Gov Tender Dashboard  ->  http://localhost:{port}\n")
    app.run(debug=debug, host="0.0.0.0", port=port, use_reloader=False)
