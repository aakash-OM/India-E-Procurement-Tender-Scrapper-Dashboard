"""
GeM Portal Bid Scraper  —  v6
==============================
CHANGES from v5:
  ✅ FIX — Organisation Name, Item Category, Quantity, Estimated Value no longer N/A
           (PDF uses SPACE not COLON between label and value — regex corrected)
  ✅ NEW — "Bid Start Date" column added before "Bid End Date"
  ✅ REMOVED — EMD Amount, Contract Period, Bid Type, MSE Preference columns
  ✅ Column order: BidNo | Org | Dept | State | Category | Qty | Value | StartDate | EndDate | Keyword | URL | ScrapedOn

Requirements:
    pip install selenium openpyxl webdriver-manager beautifulsoup4 lxml pypdf requests

Usage:
    python gem_bid_scraper.py
"""

import time, re, sys, io
from datetime import datetime
from pathlib import Path

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────

KEYWORDS = [
    # Fittings & Connectors
    "copper brass fitting transformer", "LV HV bushing transformer",
    "bimetallic connector transformer", "aluminium connector transformer",
    "connecting lug transformer", "copper lug transformer",
    "epoxy bushing transformer", "bushing flange transformer",
    "insulating bushing transformer",
    # Tap Changers
    "tap changer transformer", "tap switch 33KV", "tap switch 11KV",
    "linear tap switch", "OLTC transformer",
    # Breathers & Relays
    "silica gel breather", "acrylic breather transformer", "buchholz relay",
    "WTI OTI transformer", "oil temperature indicator",
    "MOG transformer", "marshalling box transformer", "PRV transformer",
    # Tank Accessories
    "transformer tank accessories", "drain valve transformer",
    "wheel valve transformer", "flange valve transformer",
    "air release plug transformer", "oil level gauge transformer",
    "radiator valve transformer", "MS nipple transformer",
    # Insulating Materials
    "press board transformer", "craft paper transformer",
    "crepe paper transformer", "cotton tape transformer",
    "glass sleeve transformer", "empire sleeve transformer",
    "bakelite tube transformer", "cork washer transformer",
    "nomex paper transformer", "insulation material transformer",
    # Copper & Aluminium
    "copper busbar", "aluminium busbar", "copper jumper",
    "copper thimble", "aluminium thimble", "copper sheet transformer",
    # Transmission Line
    "JAW clamp transmission", "PG clamp transmission",
    "transmission line fittings", "arcing horn transformer",
    # CT & PT
    "CT pocket transformer", "WTI pocket", "epoxy terminal transformer",
    "current transformer parts",
    # Fabrication
    "lifting hook transformer", "MS hardware transformer",
    "bush nut transformer", "soldering wire transformer",
    # Broad
    "power transformer parts", "distribution transformer accessories",
    "transformer metal parts", "transformer spares", "transformer accessories",
]

# ── SECTOR FILTER ──────────────────────────────────────────────────────────────
# Only save bids from these organisations.
# Set to EMPTY LIST  []  to get ALL bids without filtering.
TARGET_ORGANISATIONS = [
    "BHEL", "Bharat Heavy Electricals",
    "NTPC", "National Thermal Power",
    "UPPTCL", "UPPCL", "Uttar Pradesh Power Corporation",
    "KESCO", "Kanpur Electricity Supply",
    "BPCL", "Bharat Petroleum",
    "NHAI", "National Highways Authority",
    "Tata Power",
    "PGCIL", "Power Grid Corporation",
    "RVNL", "Rail Vikas Nigam",
    "IOCL", "Indian Oil",
    "Railways", "Indian Railways",
    "Power Department", "Energy Department",
    "Electricity Board", "Vidyut", "Transco", "Genco",
    "DISCOMS", "DISCOM",
]

OUTPUT_FILE           = "GeM_Bids_Master 3.xlsx"
MAX_PAGES_PER_KEYWORD = 5
HEADLESS              = True    # Set False to watch browser live
AJAX_WAIT             = 7       # Seconds after clicking Search
PAGE_NAV_WAIT         = 5       # Seconds after clicking Next page
DOWNLOAD_BID_PDF      = True    # Download each bid PDF for full details
PDF_TIMEOUT           = 20      # Seconds timeout per PDF download

# ───────────────────────────────────────────────────────────────────────────────
# ── COLUMNS (v6 — cleaner set, Start Date added) ──────────────────────────────
HEADERS = [
    "Bid Number",
    "Organisation Name",
    "Department",
    "State / Ministry",
    "Item Category",
    "Quantity",
    "Estimated Value (₹)",
    "Bid Start Date",       # ← NEW
    "Bid End Date",
    "Keyword Used",
    "Bid URL",
    "Scraped On",
]

GEM_BASE = "https://bidplus.gem.gov.in"
COL_HDR  = "1F4E79"
COL_ALT  = "DEEAF1"
COL_WHT  = "FFFFFF"
COL_SUM  = "2E75B6"
COL_GRN  = "E2EFDA"


# ─── SESSION-DEAD SIGNAL ───────────────────────────────────────────────────────

class SessionDead(Exception):
    pass


# ─── BROWSER ───────────────────────────────────────────────────────────────────

def make_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    o = Options()
    if HEADLESS:
        o.add_argument("--headless=new")
    o.add_argument("--no-sandbox")
    o.add_argument("--disable-dev-shm-usage")
    o.add_argument("--disable-gpu")
    o.add_argument("--window-size=1920,1080")
    o.add_argument("--disable-blink-features=AutomationControlled")
    o.add_experimental_option("excludeSwitches", ["enable-automation"])
    o.add_experimental_option("useAutomationExtension", False)
    o.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    drv = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=o)
    drv.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    return drv


# ─── SECTOR FILTER ─────────────────────────────────────────────────────────────

def is_target_org(org_text):
    if not TARGET_ORGANISATIONS:
        return True
    if not org_text or org_text == "N/A":
        return False
    org_upper = org_text.upper()
    return any(t.upper() in org_upper for t in TARGET_ORGANISATIONS)


# ─── SCRAPING ──────────────────────────────────────────────────────────────────

def search_keyword(driver, keyword, max_pages):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from bs4 import BeautifulSoup

    bids = []
    print(f"\n  Searching: '{keyword}'")

    try:
        _ = driver.title
    except Exception:
        raise SessionDead()

    try:
        driver.get(f"{GEM_BASE}/all-bids")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "searchBid")))
        time.sleep(2)

        box = driver.find_element(By.ID, "searchBid")
        box.clear()
        box.send_keys(keyword)
        time.sleep(0.5)
        driver.find_element(By.ID, "searchBidRA").click()
        time.sleep(AJAX_WAIT)

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a.bid_no_hover"))
            )
        except Exception:
            print("   No results.")
            return bids

        for page in range(1, max_pages + 1):
            print(f"   Page {page}...", end=" ", flush=True)
            soup   = BeautifulSoup(driver.page_source, "lxml")
            parsed = parse_bid_cards(soup, keyword)
            bids.extend(parsed)
            print(f"{len(parsed)} bids")
            if page < max_pages:
                if not go_next_page(driver, page):
                    break
                time.sleep(PAGE_NAV_WAIT)

    except SessionDead:
        raise
    except Exception as e:
        print(f"   Error: {e}")

    return bids


def go_next_page(driver, current_page):
    from selenium.webdriver.common.by import By
    try:
        links  = driver.find_elements(By.CSS_SELECTOR, "#light-pagination a, .pagination2 a")
        target = str(current_page + 1)
        for lnk in links:
            if lnk.text.strip() == target:
                lnk.click(); return True
        for lnk in links:
            if "»" in lnk.text or lnk.get_attribute("rel") == "next":
                lnk.click(); return True
    except Exception:
        pass
    return False


# ─── BID CARD PARSING ──────────────────────────────────────────────────────────

def parse_bid_cards(soup, keyword):
    results   = []
    container = soup.find("div", id="bidCard")
    if not container:
        return results
    for card in (container.find_all("div", class_="card") or []):
        bid = parse_single_card(card, keyword)
        if bid:
            results.append(bid)
    return results


def parse_single_card(card, keyword):
    try:
        bid_link = card.find("a", class_="bid_no_hover")
        if not bid_link:
            return None
        bid_no = bid_link.get_text(strip=True)
        if not re.search(r"GEM/\d{4}/", bid_no, re.I):
            return None

        href = bid_link.get("href", "")
        if href.startswith("/"):
            bid_url = GEM_BASE + href
        elif href.startswith("http"):
            bid_url = href
        else:
            bid_url = GEM_BASE + "/" + href.lstrip("/")

        f = extract_card_fields(card)

        return {
            "Bid Number":          bid_no,
            "Organisation Name":   "N/A",             # always from PDF
            "Department":          f.get("dept",       "N/A"),
            "State / Ministry":    f.get("state",      "N/A"),
            "Item Category":       "N/A",             # always from PDF
            "Quantity":            f.get("quantity",   "N/A"),
            "Estimated Value (₹)": "N/A",             # always from PDF
            "Bid Start Date":      f.get("start_date", "N/A"),
            "Bid End Date":        f.get("end_date",   "N/A"),
            "Keyword Used":        keyword,
            "Bid URL":             bid_url,
            "Scraped On":          datetime.now().strftime("%d-%m-%Y %H:%M"),
        }
    except Exception:
        return None


def extract_card_fields(card):
    """
    Confirmed GeM card text format (from live data):
    "[Qty] Department Name And Address: [Ministry] [Dept] Start Date: [X] End Date: [Y]"
    """
    result = {}
    text   = card.get_text(separator=" ", strip=True)

    # Quantity — leading number before "Department Name And Address:"
    m = re.match(r'^([\d,\.]+)\s+Department Name And Address:', text, re.I)
    if m:
        result['quantity'] = m.group(1).replace(',', '').strip()

    # Ministry/State + Department block
    m = re.search(r'Department Name And Address:\s*(.+?)\s+Start Date:', text, re.I | re.S)
    if m:
        dept_block = re.sub(r'\s*\bNA\b\s*$', '', m.group(1).strip()).strip()
        ministry, dept = _split_ministry_dept(dept_block)
        result['state'] = ministry
        if dept:
            result['dept'] = dept

    # Start Date
    m = re.search(r'Start Date:\s*(\d{2}-\d{2}-\d{4}\s+\d+:\d+\s+(?:AM|PM))', text, re.I)
    if m:
        result['start_date'] = m.group(1).strip()

    # End Date
    m = re.search(r'End Date:\s*(\d{2}-\d{2}-\d{4}\s+\d+:\d+\s+(?:AM|PM))', text, re.I)
    if m:
        result['end_date'] = m.group(1).strip()

    return result


def _split_ministry_dept(dept_block):
    """Split 'Ministry of X Department of Y' into (ministry, dept_or_None)."""
    if dept_block.startswith("PMO "):
        return ("PMO", dept_block[4:].strip())
    m = re.match(
        r'^(Ministry of [^,]+?)\s+'
        r'((?:Department|Directorate|Office|Authority|Board|Commission|'
        r'Council|Secretariat|Division).+)$', dept_block, re.I)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    m = re.match(
        r'^(Ministry of [^,]+?)\s+([A-Z]{2}[A-Z\s]+(?:Limited|Ltd|Corporation|Board).*)$',
        dept_block)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    m = re.match(
        r'^(Ministry of \w+(?:\s+\w+)?)\s+((?:Indian|Central|National|State)\s+\w.+)$',
        dept_block, re.I)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    return (dept_block, None)


# ─── PDF DOWNLOAD & PARSING ────────────────────────────────────────────────────

def fetch_pdf_bytes(bid_url):
    try:
        import requests
        r = requests.get(bid_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": GEM_BASE,
        }, timeout=PDF_TIMEOUT, stream=True)
        if r.status_code == 200 and "pdf" in r.headers.get("content-type","").lower():
            return r.content
    except Exception:
        pass
    return None


def parse_pdf_fields(pdf_bytes):
    """
    Extract fields from GeM bid PDF.
    
    KEY FIX: GeM PDFs use SPACE (not colon) between label and value:
      "Organisation Name Uttar Pradesh Power Corporation Limited (uppcl)"
      "Item Category Electric Cabling Service - XLPE..."
      "Estimated Bid Value 891000"
    
    Old regex used [:\\-] which never matched → all N/A.
    New regex: label + whitespace + value (stop at | or end of line).
    """
    if not pdf_bytes:
        return {}
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return {}

    raw_text = ""
    for page in reader.pages:
        try:
            raw_text += (page.extract_text() or "") + "\n"
        except Exception:
            continue

    if not raw_text.strip():
        return {}

    # Keep only English lines (filter bilingual Hindi/English PDF)
    clean_lines = []
    for line in raw_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        ascii_count = sum(1 for c in line if ord(c) < 128)
        if ascii_count / len(line) > 0.5:
            clean_lines.append(line)
    text = " | ".join(clean_lines)

    if not text:
        return {}

    def get(patterns):
        for pat in patterns:
            try:
                m = re.search(pat, text, re.IGNORECASE)
                if m:
                    val = m.group(1).strip().rstrip("|").strip()
                    # Remove duplicated words (bilingual PDF artifact)
                    val = re.sub(r'\b(\w+)\s+\1\b', r'\1', val)
                    if val and len(val) < 400:
                        return val
            except Exception:
                continue
        return "N/A"

    return {
        # ── FIX: use \s+ (space) not [:\-] ───────────────────────────────
        "Organisation Name": get([
            r"Organisation Name\s+([^\|]{5,120}?)(?=\s*\||\s*$)",
        ]),
        "Department": get([
            r"Department Name\s+([^\|]{5,100}?)(?=\s*\||\s*$)",
        ]),
        "State / Ministry": get([
            r"Ministry/State Name\s+([^\|]{3,80}?)(?=\s*\||\s*$)",
            r"Ministry\s+of\s+\w[^\|]{2,60}(?=\s*\|)",
        ]),
        "Item Category": get([
            r"Item Category\s+([^\|]{10,300}?)(?=\s*\||\s*$)",
            r"/Item Category\s+([^\|]{10,300}?)(?=\s*\||\s*$)",
        ]),
        "Quantity": get([
            # From consignee table: "600 N/A" at the very end
            r"(\d[\d,]*)\s+N/A\s*(?:\||$)",
            # Explicit label
            r"(?:Quantity|Qty)\s+(\d[\d,]*)",
        ]),
        "Estimated Value (₹)": get([
            r"Estimated Bid Value\s+([\d,\.]+)",
            r"Estimated Bid Value\s*[:\-]?\s*([\d,\.]+)",
        ]),
        "Bid Start Date": get([
            r"Bid Start Date/Time\s+(\d{2}-\d{2}-\d{4}[^\|]{0,20}?)(?=\s*\||\s*$)",
            r"Bid Start Date\s+(\d{2}-\d{2}-\d{4}[^\|]{0,20}?)(?=\s*\||\s*$)",
        ]),
        "Bid End Date": get([
            r"Bid End Date/Time\s+(\d{2}-\d{2}-\d{4}[^\|]{0,20}?)(?=\s*\||\s*$)",
            r"Bid End Date\s+(\d{2}-\d{2}-\d{4}[^\|]{0,20}?)(?=\s*\||\s*$)",
        ]),
    }


def enrich_bid_with_pdf(bid):
    """Download bid PDF and fill in all N/A fields with real values from PDF."""
    if not DOWNLOAD_BID_PDF:
        return bid
    bid_url = bid.get("Bid URL", "")
    if not bid_url or bid_url == "N/A":
        return bid
    try:
        pdf_bytes = fetch_pdf_bytes(bid_url)
        if not pdf_bytes:
            return bid
        pdf_fields = parse_pdf_fields(pdf_bytes)
        for field, val in pdf_fields.items():
            if val and val != "N/A":
                bid[field] = val
    except Exception:
        pass
    return bid


# ─── EXCEL ─────────────────────────────────────────────────────────────────────

def get_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin")
    return {
        "hf": Font(bold=True, color="FFFFFF", name="Arial", size=10),
        "df": Font(name="Arial", size=9),
        "bf": Font(bold=True, name="Arial", size=10),
        "ce": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "le": Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "br": Border(left=thin, right=thin, top=thin, bottom=thin),
        "fh": PatternFill("solid", fgColor=COL_HDR),
        "fa": PatternFill("solid", fgColor=COL_ALT),
        "fw": PatternFill("solid", fgColor=COL_WHT),
        "fs": PatternFill("solid", fgColor=COL_SUM),
        "ft": PatternFill("solid", fgColor="BDD7EE"),
        "fg": PatternFill("solid", fgColor=COL_GRN),
    }


def write_headers(ws, s):
    from openpyxl.utils import get_column_letter
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font=s["hf"]; c.fill=s["fh"]; c.alignment=s["ce"]; c.border=s["br"]
    ws.row_dimensions[1].height = 35
    col_widths = [22, 40, 32, 24, 50, 12, 20, 20, 20, 22, 45, 18]
    for ci, w in enumerate(col_widths[:len(HEADERS)], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def append_bids(ws, bids, s):
    from openpyxl.styles import PatternFill
    existing = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    added = 0
    for bid in bids:
        bn = str(bid["Bid Number"]).strip()
        if bn in existing:
            continue
        nr  = ws.max_row + 1
        org = bid.get("Organisation Name", "")
        fl  = s["fg"] if is_target_org(org) else (
              s["fa"] if nr % 2 == 0 else PatternFill("solid", fgColor=COL_WHT))
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(nr, ci, bid.get(h, "N/A"))
            c.font=s["df"]; c.fill=fl; c.border=s["br"]
            c.alignment = s["le"] if ci in (2, 3, 4, 5, 11) else s["ce"]
        existing.add(bn)
        added += 1
    return added


def rebuild_summary(wb, s, run_stats):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    ws["A1"] = "GeM Bids — Monthly Summary"
    ws["A1"].font = Font(bold=True, name="Arial", size=14, color=COL_HDR)
    ws["A2"] = f"Last updated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    filt = (f"Sector filter: {', '.join(TARGET_ORGANISATIONS[:5])}..."
            if TARGET_ORGANISATIONS else "Sector filter: ALL")
    ws["A3"] = filt
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells("A1:F1"); ws.merge_cells("A2:F2"); ws.merge_cells("A3:F3")

    for ci, h in enumerate(["Month","Total Bids","Sector Matched","Added This Run","Keywords Used","Sheet"], 1):
        c = ws.cell(5, ci, h)
        c.font=s["hf"]; c.fill=s["fs"]; c.alignment=s["ce"]; c.border=s["br"]

    sheets = [n for n in wb.sheetnames if n != "Summary"]
    grand  = 0
    for ri, name in enumerate(sheets, 6):
        wm    = wb[name]
        count = max(wm.max_row - 1, 0)
        # Keyword is col 10 (index 9)
        kws   = {str(r[9]) for r in wm.iter_rows(min_row=2, values_only=True) if r[9]}
        sec   = sum(1 for r in wm.iter_rows(min_row=2, values_only=True)
                    if r[1] and is_target_org(str(r[1])))
        fl    = s["fa"] if ri % 2 == 0 else PatternFill("solid", fgColor=COL_WHT)
        for ci, v in enumerate([name, count, sec, run_stats.get(name,0),
                                 ", ".join(sorted(kws)) or "—", f"→ '{name}'"], 1):
            c = ws.cell(ri, ci, v)
            c.font=s["df"]; c.fill=fl; c.border=s["br"]
            c.alignment = s["le"] if ci == 5 else s["ce"]
        grand += count

    tr = 6 + len(sheets)
    for ci, v in enumerate(["GRAND TOTAL", grand, "", "", "", ""], 1):
        c = ws.cell(tr, ci, v)
        c.font=s["bf"]; c.fill=s["ft"]; c.border=s["br"]; c.alignment=s["ce"]

    for ci, w in enumerate([18, 14, 16, 16, 50, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames)-1))


def save_excel(bids, output_path, run_stats):
    from openpyxl import Workbook, load_workbook
    month = datetime.now().strftime("%B %Y")
    s     = get_styles()

    if Path(output_path).exists():
        print(f"\n  Opening existing file : {output_path}")
        wb = load_workbook(output_path)
        if month in wb.sheetnames:
            wm = wb[month]
            # Detect old schema (15 cols) vs new schema (12 cols) and wipe if needed
            old_headers = [wm.cell(1, c).value for c in range(1, wm.max_column+1)]
            if "EMD Amount (₹)" in old_headers or wm.max_column != len(HEADERS):
                print(f"  ⚠️  Old column schema detected in '{month}' — rebuilding sheet...")
                del wb[month]
    else:
        print(f"\n  Creating new file     : {output_path}")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if month not in wb.sheetnames:
        ws = wb.create_sheet(month)
        write_headers(ws, s)
        print(f"  New sheet created     : '{month}'")
    else:
        ws = wb[month]
        print(f"  Sheet found           : '{month}'")

    added   = append_bids(ws, bids, s)
    skipped = len(bids) - added
    run_stats[month] = added
    print(f"  New bids added        : {added}")
    print(f"  Duplicates skipped    : {skipped}")
    rebuild_summary(wb, s, run_stats)
    wb.save(output_path)
    print(f"  Saved                 : {output_path}")
    return added


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    now = datetime.now()
    print("=" * 62)
    print("  GeM Portal Bid Scraper  v6")
    print("=" * 62)
    print(f"  Output file  : {OUTPUT_FILE}")
    print(f"  Sheet        : '{now.strftime('%B %Y')}'")
    print(f"  PDF parsing  : {'ON' if DOWNLOAD_BID_PDF else 'OFF'}")
    print(f"  Sector filter: {', '.join(TARGET_ORGANISATIONS[:3])}..." if TARGET_ORGANISATIONS else "  Sector filter: ALL")
    print(f"  Columns      : {len(HEADERS)}  (EMD/ContractPeriod/BidType/MSE removed; Start Date added)")
    print("=" * 62)

    for pkg in ["selenium","openpyxl","webdriver_manager","bs4","lxml","pypdf","requests"]:
        try:
            __import__(pkg)
        except ImportError:
            print(f"\n  ❌ Missing: {pkg}")
            print("  Run: pip install selenium openpyxl webdriver-manager beautifulsoup4 lxml pypdf requests")
            sys.exit(1)

    all_bids, driver, kw_index = [], None, 0

    while kw_index < len(KEYWORDS):
        kw = KEYWORDS[kw_index]
        if driver is None:
            print("\n  Starting browser...")
            driver = make_driver()
        try:
            bids = search_keyword(driver, kw, MAX_PAGES_PER_KEYWORD)
            all_bids.extend(bids)
            if bids:
                print(f"  ✓ '{kw}': {len(bids)} bids")
            kw_index += 1
            time.sleep(1)
        except SessionDead:
            print(f"\n  🔄 Browser crashed. Restarting...")
            try: driver.quit()
            except Exception: pass
            driver = None
            time.sleep(3)
        except KeyboardInterrupt:
            print("\n  Interrupted. Saving data collected so far...")
            break
        except Exception as e:
            print(f"\n  Error on '{kw}': {e}")
            kw_index += 1

    if driver:
        try: driver.quit()
        except Exception: pass

    # Deduplicate
    seen, unique = set(), []
    for b in all_bids:
        if b["Bid Number"] not in seen:
            seen.add(b["Bid Number"]); unique.append(b)
    print(f"\n  Total unique bids scraped : {len(unique)}")

    if not unique:
        print("  ⚠️  No bids. Try HEADLESS=False or increase AJAX_WAIT.")
        return

    # PDF enrichment
    if DOWNLOAD_BID_PDF:
        print("  Downloading PDFs for Organisation Name, Item Category, Value, Quantity...")
        for i, bid in enumerate(unique):
            pct = int((i+1)/len(unique)*100)
            sys.stdout.write(f"\r  PDF {i+1}/{len(unique)} ({pct}%)  {bid['Bid Number'][:28]:<28}")
            sys.stdout.flush()
            try:
                unique[i] = enrich_bid_with_pdf(bid)
            except Exception:
                pass
            time.sleep(0.3)
        print(f"\r  PDF enrichment complete — {len(unique)} bids processed.          ")

    # Sector filter
    if TARGET_ORGANISATIONS:
        matched  = [b for b in unique if is_target_org(b.get("Organisation Name",""))]
        excluded = len(unique) - len(matched)
        print(f"  Sector filter  : {len(matched)} matched / {excluded} excluded")
        final_bids = matched if matched else unique  # fallback: save all if none match
        if not matched:
            print("  ⚠️  No sector matches — saving all bids to avoid data loss.")
    else:
        final_bids = unique

    run_stats = {}
    added = save_excel(final_bids, OUTPUT_FILE, run_stats)

    print("\n" + "=" * 62)
    print(f"  ✅ Done!  {added} bids added → '{now.strftime('%B %Y')}' tab")
    print(f"  📁 File  : {OUTPUT_FILE}")
    print(f"  🟢 Green = sector-matched  |  Blue = others")
    print("=" * 62)


if __name__ == "__main__":
    main()
