"""
GeM Portal Bid Scraper  —  v8  (Parallel PDF Downloads)
=====================================================
NEW in v7:
  ✅ Smart pipeline — PDFs downloaded ONLY for bids that pass ALL pre-checks:
       Step 1: Deduplicate scraped bids within this run
       Step 2: Card pre-filter  — eliminate non-power-sector by State/Dept text
       Step 3: Excel duplicate check — skip bids already saved
       Step 4: Download PDFs only for remaining candidates (~90% fewer downloads)
       Step 5: Final org-name filter using PDF Organisation Name
       Step 6: Save to Excel
  ✅ Auto-retry save if Excel file is open (PermissionError)
  ✅ Browser auto-restart on crash
  ✅ Monthly sheets, duplicate protection, green highlights for sector matches

Requirements:
    pip install selenium openpyxl webdriver-manager beautifulsoup4 lxml pypdf requests
    (No new packages needed for parallelism — uses Python stdlib ThreadPoolExecutor)
"""

import time, re, sys, io
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────

KEYWORDS = [
    # Fittings & Connectors
    "copper brass fitting transformer", "LV HV bushing transformer", "HV/LV Brass Studs",
    "bimetallic connector transformer", "aluminium connector transformer",
    "connecting lug transformer", "copper lug transformer", "Alluminium Lugs",
    "epoxy bushing transformer", "bushing flange transformer", "Rolled Copper"
    "insulating bushing transformer", "MS Earthing Rod", "LT Palm connectors",
    "Copper", "Brass", "Brass Terminal Connectors", "Terminal Connectors", "Lead Connectors",
    "Transformer Bushing Connector", "brass Stud", "Copper Stud", "Brass/AL Stud"
    # Sheets
    "copper Alloy sheet", "Copper Sheet", "Steel Sheet", "Bakelite Sheet",
    "Fuse connecting Steel Sheet", "Rubberized cork sheet", "GI Sheet",
    "Copper Rods", "Copper Bars", "solid press board sheet", "tfr Cork Sheet",
    # Tap Changers
    "tap changer transformer", "tap switch 33KV", "tap switch 11KV",
    "linear tap switch", "OLTC transformer",
    # Breathers & Relays
    "silica gel breather", "Silica Gel", "Silica Gel Crystal",
    "acrylic breather transformer", "buchholz relay",
    "WTI OTI transformer", "oil temperature indicator", "Breathers",
    "Breathers for OLTC", "MOG transformer", "marshalling box transformer",
    "PRV transformer", "SMC Box", "Single Phase Distribution Box",
    "Three Phase Distribution Box", "Breather Assembly",
    # Tank Accessories
    "transformer tank accessories", "drain valve transformer",
    "wheel valve transformer", "flange valve transformer",
    "air release plug transformer", "oil level gauge transformer",
    "radiator valve transformer", "MS nipple transformer",
    # Insulating Materials
    "press board transformer", "craft paper transformer", "LT HT Insulator",
    "crepe paper transformer", "cotton tape transformer", "LT Spacer",
    "glass sleeve transformer", "empire sleeve transformer", "bakelite tube",
    "bakelite tube transformer", "cork washer transformer",
    "nomex paper transformer", "insulation material transformer",
    "Insulated Synthetic Mat",
    # Copper & Aluminium
    "copper busbar", "aluminium busbar", "copper jumper",
    "copper thimble", "aluminium thimble", "copper sheet transformer",
    # Transmission Line
    "JAW clamp transmission", "PG clamp transmission", "BI-Metallic",
    "BI-Metallic VCB CLAMP", "transmission line fittings", "arcing horn transformer",
    "HT BI-Metallic CLAMP", "Copper Z Clamp", "PG Clamp", "CT Clamp",
    "PT Clamp", "substation clamps",
    # CT & PT
    "CT pocket transformer", "CT", "WTI pocket", "epoxy terminal transformer",
    "current transformer parts", "CTPT", "CTs", "PTs",
    # Fabrication
    "lifting hook transformer", "MS hardware transformer", "Nut", "Bolt",
    "bush nut", "bush nut transformer", "soldering wire transformer",
    "Fuse Wire", "tinned copper fuse wire",
    # Broad
    "power transformer parts", "distribution transformer accessories",
    "Metal strips", "transformer metal parts", "transformer spares",
    "transformer accessories",
    # Others
    "Viberation Damper", "zebra condutor", "Twin Mooseconductor", "bolt fitted",
    "LT Fuse", "tailless Fuse", "HRC Fuse", "Sweated Assembly", "VCB",
    "Distribution Transformer", "earthing arrangement", "nut bolt",
    "HV Porcelain Bushing 12mm", "Conservator Gauge", "HV Porcelain Bushing",
    "Brass Hex Check Nut", "Brass Plain Washer", "Bushing Metal Parts",
    "Earthing Metal Parts", "Metal Parts", "Bushing Rods",
    "Cast Iron Butterfly", "Butterfly Valves", "Throttle Valve",
    "Gutka(Spacers)", "Wedges", "Vacuum Circuit Breaker", "Lightening Arrestor",
    "Surge Arrestors", "Nuts", "Brass Flat washer", "Ring Quad Tension Strings",
    "double break isolator", "copper strip finger", "Horn Sleeves",
    "Tfr Conical washer", "Tfr cork washer", "Glass Sleeves",
    "Neoprene Oil Seals", "HT/LT",
]

# ── SECTOR FILTER ──────────────────────────────────────────────────────────────
# Final filter: only save bids whose PDF Organisation Name matches these.
# Set to EMPTY LIST [] to save ALL bids.
TARGET_ORGANISATIONS = [
    "BHEL", "Bharat Heavy Electricals", "Department of Heavy Industry",
    "NTPC", "National Thermal Power", "NHPC",
    "NHAI", "National Highways Authority",
    "Uttar Pradesh Power Corporation",
    "UP Power Transmission corp. ltd",
    "KESCO", "Kanpur Electricity Supply", "Kanpur Electricity Supply Company",
    "PGCIL", "Power Grid Corporation", "Power Grid Corporation of India Limited",
    "RVNL", "Rail Vikas Nigam",
    "UPPCL", "Uttar Pradesh Power Corporation Limited",
    "UPPTCL", "Uttar Pradesh Power Transmission Corporation Limited",
    "UPRVUNL", "Uttar Pradesh Rajya Vidyut Utpadan Nigam Limited",
    "UPJVNL", "Uttar Pradesh Jal Vidyut Nigam Limited",
    "PVVNL", "Paschimanchal Vidyut Vitran Nigam Limited",
    "MVVNL", "Madhyanchal Vidyut Vitran Nigam Limited",
    "DVVNL", "Dakshinanchal Vidyut Vitran Nigam Limited",
    "PuVVNL", "Purvanchal Vidyut Vitran Nigam Limited",
    "NPCL", "Noida Power Company Limited",
    "DTL", "Delhi Transco Limited", "Indraprastha Power Generation Co. Ltd.",
    "BSES Rajdhani", "BSES Rajdhani Power Limited",
    "BSES Yamuna", "BSES Yamuna Power Limited",
    "TPDDL", "Tata Power Delhi Distribution Limited",
    "MSEB", "Maharashtra State Electricity Board",
    "MAHAGENCO", "Maharashtra State Power Generation Company Limited",
    "MAHATRANSCO", "Maharashtra State Electricity Transmission Company Limited",
    "MSEDCL", "Maharashtra State Electricity Distribution Company Limited",
    "BEST", "Brihanmumbai Electric Supply and Transport",
    "Tata Power", "The Tata Power Company Limited",
    "GUVNL", "Gujarat Urja Vikas Nigam Limited",
    "GSECL", "Gujarat State Electricity Corporation Limited",
    "GETCO", "Gujarat Energy Transmission Corporation Limited",
    "UGVCL", "Uttar Gujarat Vij Company Limited",
    "DGVCL", "Dakshin Gujarat Vij Company Limited",
    "MGVCL", "Madhya Gujarat Vij Company Limited",
    "PGVCL", "Paschim Gujarat Vij Company Limited",
    "MPPGCL", "Madhya Pradesh Power Generating Company Limited",
    "MPPTCL", "Madhya Pradesh Power Transmission Company Limited", "MPPKVVCL",
    "MPMKVVCL", "Madhya Pradesh Madhya Kshetra Vidyut Vitaran Company Limited",
    "PSPCL", "Punjab State Power Corporation Limited",
    "PSTCL", "Punjab State Transmission Corporation Limited",
    "HPGCL", "Haryana Power Generation Corporation Limited",
    "HVPNL", "Haryana Vidyut Prasaran Nigam Limited",
    "DHBVN", "Dakshin Haryana Bijli Vitran Nigam",
    "UHBVN", "Uttar Haryana Bijli Vitran Nigam",
    "KPCL", "Karnataka Power Corporation Limited",
    "KPTCL", "Karnataka Power Transmission Corporation Limited",
    "BESCOM", "Bangalore Electricity Supply Company Limited",
    "MESCOM", "Mangalore Electricity Supply Company Limited",
    "HESCOM", "Hubli Electricity Supply Company Limited",
    "GESCOM", "Gulbarga Electricity Supply Company Limited",
    "CESC", "Chamundeshwari Electricity Supply Corporation Limited",
    "CESC", "Calcutta Electric Supply Corporation Limited",
    "WBPDCL", "West Bengal Power Development Corporation Limited",
    "WBSETCL", "West Bengal State Electricity Transmission Company Limited",
    "WBSEDCL", "West Bengal State Electricity Distribution Company Limited",
    "JKSPDC", "Jammu and Kashmir State Power Development Corporation",
    "JKPTCL", "Jammu and Kashmir Power Transmission Corporation Limited",
    "JKPDD", "Jammu and Kashmir Power Distribution Department",
    "HPSEBL", "Himachal Pradesh State Electricity Board Limited",
    "HPPTCL", "Himachal Pradesh Power Transmission Corporation Limited",
    "SJVN", "Satluj Jal Vidyut Nigam Limited",
    "UJVNL", "Uttarakhand Jal Vidyut Nigam Limited",
    "PTCUL", "Power Transmission Corporation of Uttarakhand Limited",
    "UPCL", "Uttarakhand Power Corporation Limited",
    "TVNL", "Tenughat Vidyut Nigam Limited",
    "JBVNL", "Jharkhand Bijli Vitran Nigam Limited",
    "JSEB", "Jharkhand State Electricity Board",
    "CSPGCL", "Chhattisgarh State Power Generation Company Limited",
    "CSPTCL", "Chhattisgarh State Power Transmission Company Limited",
    "CSPDCL", "Chhattisgarh State Power Distribution Company Limited",
    "BSPGCL", "Bihar State Power Generation Company Limited",
    "BSPTCL", "Bihar State Power Transmission Company Limited",
    "NBPDCL", "North Bihar Power Distribution Company Limited",
    "SBPDCL", "South Bihar Power Distribution Company Limited",
    "Railways", "Indian Railways",
    "Power Department", "Energy Department",
    "Electricity Board", "Vidyut", "Transco", "Genco",
    "DISCOMS", "DISCOM",
    "Ministry Of Heavy Industries and Public Enterprises",
    "Patratu Vidyut Utpadan Nigam",
]

# ── CARD PRE-FILTER KEYWORDS ───────────────────────────────────────────────────
# Broad keywords checked against card State/Ministry + Department text BEFORE
# downloading any PDF. Intentionally inclusive — catches all power/energy bids.
# False positives are fine; false negatives (missing a real match) are not.
CARD_SECTOR_KEYWORDS = [
    "Power", "Energy", "Electric", "Vidyut", "Bijli",
    "Transco", "Genco", "Discom", "Grid",
    "Transmission", "Generation", "Distribution",
    "Railways", "Railway", "NTPC", "BHEL", "NHPC", "PGCIL", "NHAI",
    "Heavy Industries",
    # State names where we have target utilities
    "Uttar Pradesh", "Maharashtra", "Gujarat", "Punjab", "Haryana",
    "Karnataka", "West Bengal", "Jharkhand", "Chhattisgarh", "Bihar",
    "Madhya Pradesh", "Himachal", "Uttarakhand", "Jammu", "Delhi",
    "Rajasthan", "Andhra Pradesh", "Telangana", "Tamil Nadu", "Kerala",
    "Odisha", "Assam", "Patratu",
]

# ── BLOCKED ITEM CATEGORIES ───────────────────────────────────────────────────
# Bids whose Item Category (from PDF) contains ANY of these words will be
# REJECTED — even if they belong to a TARGET_ORGANISATION.
# Catches irrelevant bids: taxis, repairs, cleaning, vehicles etc.
BLOCKED_KEYWORDS = [
    # Services / labour contracts
    "Repair", "Repairing", "Maintenance", "Installation", "Operation",
    "Cleaning", "Housekeeping", "House Keeping", "Sweeping", "Sanitation",
    "Facility Management", "Management Service", "Management Services",
    "Carriage", "Waiting Area", "Advertisement", " The New Indian Express", "Service","Services",
    # Vehicles / transport
    "Cab", "Taxi", "Hiring", "Vehicle", "Car Rental", "Bus Hiring",
    "SUV", "Mahindra", "Scorpio", "Bolero", "Head Light",
    "Maruti", "Suzuki", "Ertiga", "Hyundai", "Creta",
    # Other unrelated items
    "Pedestal", "Handicraft", "Double Wooden Bed", "Wood", "BLOWDOWNVALVEKIT",
    #chemicals
    "ALUMINIUM SULPHATE",
    "Paint", "Zinc Priming Paint", "Zinc Priming Paint","Hardner", "Solvent","Resin","Primer",
    "Zinc Phosphate Primer", "Zinc", "Resin", "Acid",
]

OUTPUT_FILE           = "GeM_Bids_April.xlsx"
MAX_PAGES_PER_KEYWORD = 5
HEADLESS              = False
AJAX_WAIT             = 7
PAGE_NAV_WAIT         = 5
DOWNLOAD_BID_PDF      = True
PDF_TIMEOUT           = 20
PARALLEL_PDF_WORKERS  = 10   # PDFs downloaded simultaneously (8-15 recommended)

# ───────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "Bid Number", "Organisation Name", "Department", "State / Ministry",
    "Item Category", "Quantity", "Estimated Value (₹)",
    "Bid Start Date", "Bid End Date",
    "Keyword Used", "Bid URL", "Scraped On",
]

GEM_BASE = "https://bidplus.gem.gov.in"
COL_HDR  = "1F4E79"
COL_ALT  = "DEEAF1"
COL_WHT  = "FFFFFF"
COL_SUM  = "2E75B6"
COL_GRN  = "E2EFDA"


# ─── SESSION-DEAD SIGNAL ───────────────────────────────────────────────────────

class SessionDead(Exception):
    pass


# ─── BROWSER ───────────────────────────────────────────────────────────────────

def make_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    o = Options()
    if HEADLESS:
        o.add_argument("--headless=new")
    o.add_argument("--no-sandbox")
    o.add_argument("--disable-dev-shm-usage")
    o.add_argument("--disable-gpu")
    o.add_argument("--window-size=1920,1080")
    o.add_argument("--disable-blink-features=AutomationControlled")
    o.add_experimental_option("excludeSwitches", ["enable-automation"])
    o.add_experimental_option("useAutomationExtension", False)
    o.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    drv = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=o)
    drv.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    return drv


# ─── FILTERS ───────────────────────────────────────────────────────────────────

def is_target_org(org_text):
    """Final filter: check PDF Organisation Name against TARGET_ORGANISATIONS."""
    if not TARGET_ORGANISATIONS:
        return True
    if not org_text or org_text == "N/A":
        return False
    org_upper = org_text.upper()
    return any(t.upper() in org_upper for t in TARGET_ORGANISATIONS)


def is_blocked_category(item_category):
    """
    Return True if Item Category contains a blocked keyword.
    These bids are rejected even if the org matches TARGET_ORGANISATIONS.
    """
    if not item_category or item_category == "N/A":
        return False
    cat_upper = item_category.upper()
    return any(kw.upper() in cat_upper for kw in BLOCKED_KEYWORDS)


def card_prefilter(bid):
    """
    Quick pre-filter using card data (no PDF download needed).
    Checks State/Ministry + Department text against broad sector keywords.
    Intentionally broad — no false negatives allowed.
    """
    if not TARGET_ORGANISATIONS:
        return True   # no filter active
    state = str(bid.get("State / Ministry", "") or "").upper()
    dept  = str(bid.get("Department", "") or "").upper()
    combined = state + " " + dept
    # Check exact target names first
    for t in TARGET_ORGANISATIONS:
        if t.upper() in combined:
            return True
    # Check broad sector keywords
    for kw in CARD_SECTOR_KEYWORDS:
        if kw.upper() in combined:
            return True
    return False


def get_existing_bid_numbers(output_path):
    """Read bid numbers already saved in the current month's Excel sheet."""
    existing = set()
    month = datetime.now().strftime("%B %Y")
    try:
        if not Path(output_path).exists():
            return existing
        from openpyxl import load_workbook
        wb = load_workbook(output_path, read_only=True)
        if month not in wb.sheetnames:
            wb.close()
            return existing
        ws = wb[month]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                existing.add(str(row[0]).strip())
        wb.close()
    except Exception:
        pass
    return existing


# ─── SCRAPING ──────────────────────────────────────────────────────────────────

def search_keyword(driver, keyword, max_pages):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from bs4 import BeautifulSoup

    bids = []
    print(f"\n  Searching: '{keyword}'")

    try:
        _ = driver.title
    except Exception:
        raise SessionDead()

    try:
        driver.get(f"{GEM_BASE}/all-bids")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "searchBid")))
        time.sleep(2)

        box = driver.find_element(By.ID, "searchBid")
        box.clear()
        box.send_keys(keyword)
        time.sleep(0.5)
        driver.find_element(By.ID, "searchBidRA").click()
        time.sleep(AJAX_WAIT)

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a.bid_no_hover"))
            )
        except Exception:
            print("   No results.")
            return bids

        for page in range(1, max_pages + 1):
            print(f"   Page {page}...", end=" ", flush=True)
            soup   = BeautifulSoup(driver.page_source, "lxml")
            parsed = parse_bid_cards(soup, keyword)
            bids.extend(parsed)
            print(f"{len(parsed)} bids")
            if page < max_pages:
                if not go_next_page(driver, page):
                    break
                time.sleep(PAGE_NAV_WAIT)

    except SessionDead:
        raise
    except Exception as e:
        print(f"   Error: {e}")

    return bids


def go_next_page(driver, current_page):
    from selenium.webdriver.common.by import By
    try:
        links  = driver.find_elements(By.CSS_SELECTOR, "#light-pagination a, .pagination2 a")
        target = str(current_page + 1)
        for lnk in links:
            if lnk.text.strip() == target:
                lnk.click(); return True
        for lnk in links:
            if "»" in lnk.text or lnk.get_attribute("rel") == "next":
                lnk.click(); return True
    except Exception:
        pass
    return False


# ─── BID CARD PARSING ──────────────────────────────────────────────────────────

def parse_bid_cards(soup, keyword):
    results   = []
    container = soup.find("div", id="bidCard")
    if not container:
        return results
    for card in (container.find_all("div", class_="card") or []):
        bid = parse_single_card(card, keyword)
        if bid:
            results.append(bid)
    return results


def parse_single_card(card, keyword):
    try:
        bid_link = card.find("a", class_="bid_no_hover")
        if not bid_link:
            return None
        bid_no = bid_link.get_text(strip=True)
        if not re.search(r"GEM/\d{4}/", bid_no, re.I):
            return None

        href = bid_link.get("href", "")
        if href.startswith("/"):
            bid_url = GEM_BASE + href
        elif href.startswith("http"):
            bid_url = href
        else:
            bid_url = GEM_BASE + "/" + href.lstrip("/")

        f = extract_card_fields(card)

        return {
            "Bid Number":          bid_no,
            "Organisation Name":   "N/A",
            "Department":          f.get("dept",       "N/A"),
            "State / Ministry":    f.get("state",      "N/A"),
            "Item Category":       "N/A",
            "Quantity":            f.get("quantity",   "N/A"),
            "Estimated Value (₹)": "N/A",
            "Bid Start Date":      f.get("start_date", "N/A"),
            "Bid End Date":        f.get("end_date",   "N/A"),
            "Keyword Used":        keyword,
            "Bid URL":             bid_url,
            "Scraped On":          datetime.now().strftime("%d-%m-%Y %H:%M"),
        }
    except Exception:
        return None


def extract_card_fields(card):
    result = {}
    text   = card.get_text(separator=" ", strip=True)

    m = re.match(r'^([\d,\.]+)\s+Department Name And Address:', text, re.I)
    if m:
        result['quantity'] = m.group(1).replace(',', '').strip()

    m = re.search(r'Department Name And Address:\s*(.+?)\s+Start Date:', text, re.I | re.S)
    if m:
        dept_block = re.sub(r'\s*\bNA\b\s*$', '', m.group(1).strip()).strip()
        ministry, dept = _split_ministry_dept(dept_block)
        result['state'] = ministry
        if dept:
            result['dept'] = dept

    m = re.search(r'Start Date:\s*(\d{2}-\d{2}-\d{4}\s+\d+:\d+\s+(?:AM|PM))', text, re.I)
    if m:
        result['start_date'] = m.group(1).strip()

    m = re.search(r'End Date:\s*(\d{2}-\d{2}-\d{4}\s+\d+:\d+\s+(?:AM|PM))', text, re.I)
    if m:
        result['end_date'] = m.group(1).strip()

    return result


def _split_ministry_dept(dept_block):
    if dept_block.startswith("PMO "):
        return ("PMO", dept_block[4:].strip())
    m = re.match(
        r'^(Ministry of [^,]+?)\s+'
        r'((?:Department|Directorate|Office|Authority|Board|Commission|'
        r'Council|Secretariat|Division).+)$', dept_block, re.I)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    m = re.match(
        r'^(Ministry of [^,]+?)\s+([A-Z]{2}[A-Z\s]+(?:Limited|Ltd|Corporation|Board).*)$',
        dept_block)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    m = re.match(
        r'^(Ministry of \w+(?:\s+\w+)?)\s+((?:Indian|Central|National|State)\s+\w.+)$',
        dept_block, re.I)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    return (dept_block, None)


# ─── PDF DOWNLOAD & PARSING ────────────────────────────────────────────────────

def fetch_pdf_bytes(bid_url):
    try:
        import requests
        r = requests.get(bid_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": GEM_BASE,
        }, timeout=PDF_TIMEOUT, stream=True)
        if r.status_code == 200 and "pdf" in r.headers.get("content-type", "").lower():
            return r.content
    except Exception:
        pass
    return None


def parse_pdf_fields(pdf_bytes):
    if not pdf_bytes:
        return {}
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return {}

    raw_text = ""
    for page in reader.pages:
        try:
            raw_text += (page.extract_text() or "") + "\n"
        except Exception:
            continue

    if not raw_text.strip():
        return {}

    clean_lines = []
    for line in raw_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        if sum(1 for c in line if ord(c) < 128) / len(line) > 0.5:
            clean_lines.append(line)
    text = " | ".join(clean_lines)

    if not text:
        return {}

    def get(patterns):
        for pat in patterns:
            try:
                m = re.search(pat, text, re.IGNORECASE | re.S)
                if m:
                    val = m.group(1).strip().strip("|").strip()
                    val = re.sub(r'\b(\w+)\s+\1\b', r'\1', val)
                    val = re.sub(r'\s*\|\s*', ' ', val).strip()
                    if val and 2 < len(val) < 400:
                        return val
            except Exception:
                continue
        return "N/A"

    return {
        "Organisation Name": get([r"/Organisation Name\s*\|\s*([^|]{5,120}?)\s*\|"]),
        "Department":        get([r"/Department Name\s*\|\s*([^|]{5,100}?)\s*\|"]),
        "State / Ministry":  get([r"/Ministry/State Name\s*\|\s*([^|]{3,80}?)\s*\|"]),
        "Item Category":     get([
            r"/Item Category\s*\|\s*(.*?)\s*\|?\s*/Contract Period",
            r"/Item Category\s*\|\s*([^|]{10,300}?)\s*\|",
        ]),
        "Quantity": get([
            r"Total Quantity\s*\|\s*(\d[\d,]*)",
            r"(?:कुल मात्रा/)?Total Quantity\s*\|\s*(\d[\d,]*)",
            r"(\d[\d,]*)\s*\|\s*N/A",
            r"(\d[\d,]*)\s+N/A",
        ]),
        "Estimated Value (₹)": get([
            r"/ Estimated Bid Value\s*\|\s*([\d,\.]+)",
            r"/Estimated Bid Value\s*\|\s*([\d,\.]+)",
            r"Estimated Bid Value\s*\|\s*([\d,\.]+)",
        ]),
        "Bid Start Date": get([
            r"/Bid Start Date/Time\s*\|\s*(\d{2}-\d{2}-\d{4}[^|]{0,20}?)\s*\|",
        ]),
        "Bid End Date": get([
            r"/Bid End Date/Time\s*\|\s*(\d{2}-\d{2}-\d{4}[^|]{0,20}?)\s*\|",
        ]),
    }


def enrich_bid_with_pdf(bid):
    bid_url = bid.get("Bid URL", "")
    if not bid_url or bid_url == "N/A":
        return bid
    try:
        pdf_bytes = fetch_pdf_bytes(bid_url)
        if not pdf_bytes:
            return bid
        pdf_fields = parse_pdf_fields(pdf_bytes)
        DATE_FIELDS = {"Bid Start Date", "Bid End Date"}   # always keep card dates
        for field, val in pdf_fields.items():
            if field in DATE_FIELDS:
                continue
            if val and val != "N/A":
                bid[field] = val
    except Exception:
        pass
    return bid



# ─── PARALLEL PDF ENRICHMENT ───────────────────────────────────────────────────

def _enrich_one(bid):
    """
    Single-bid worker executed inside a thread-pool thread.
    Identical logic to enrich_bid_with_pdf() — kept separate so the
    original function still works for any one-off calls elsewhere.
    Thread-safe: reads only its own `bid` dict, no shared mutable state.
    """
    bid_url = bid.get("Bid URL", "")
    if not bid_url or bid_url == "N/A":
        return bid
    try:
        pdf_bytes = fetch_pdf_bytes(bid_url)
        if not pdf_bytes:
            return bid
        pdf_fields = parse_pdf_fields(pdf_bytes)
        DATE_FIELDS = {"Bid Start Date", "Bid End Date"}
        for field, val in pdf_fields.items():
            if field in DATE_FIELDS:
                continue
            if val and val != "N/A":
                bid[field] = val
    except Exception:
        pass
    return bid


def enrich_bids_parallel(bids, max_workers=PARALLEL_PDF_WORKERS):
    """
    Download + parse PDFs for all bids in parallel using a thread pool.

    WHY ThreadPoolExecutor AND NOT asyncio:
      • requests (HTTP) and pypdf (parsing) are both blocking/CPU libs —
        they cannot be used with asyncio without rewriting everything.
      • Threads give true I/O parallelism here with zero library changes.
      • 10 threads ≈ 10× speedup: 56 min → ~5 min for 785 PDFs.

    Results list preserves original order even though futures complete
    out-of-order. Progress counter updates in real time.
    """
    total   = len(bids)
    results = [None] * total          # pre-allocated — preserves order
    done    = [0]                     # list so the closure can mutate it

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_idx = {
            pool.submit(_enrich_one, bid): i
            for i, bid in enumerate(bids)
        }
        for future in as_completed(future_to_idx):
            idx          = future_to_idx[future]
            results[idx] = future.result()
            done[0]     += 1
            pct          = int(done[0] / total * 100)
            bid_no       = bids[idx].get("Bid Number", "")[:28]
            sys.stdout.write(
                f"\r  PDF {done[0]}/{total} ({pct}%)  {bid_no:<28}"
            )
            sys.stdout.flush()

    print(f"\r  PDF enrichment complete — {total} bids ({max_workers} parallel workers).          ")
    return results


# ─── EXCEL ─────────────────────────────────────────────────────────────────────

def get_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin")
    return {
        "hf": Font(bold=True, color="FFFFFF", name="Arial", size=10),
        "df": Font(name="Arial", size=9),
        "bf": Font(bold=True, name="Arial", size=10),
        "ce": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "le": Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "br": Border(left=thin, right=thin, top=thin, bottom=thin),
        "fh": PatternFill("solid", fgColor=COL_HDR),
        "fa": PatternFill("solid", fgColor=COL_ALT),
        "fw": PatternFill("solid", fgColor=COL_WHT),
        "fs": PatternFill("solid", fgColor=COL_SUM),
        "ft": PatternFill("solid", fgColor="BDD7EE"),
        "fg": PatternFill("solid", fgColor=COL_GRN),
    }


def write_headers(ws, s):
    from openpyxl.utils import get_column_letter
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font=s["hf"]; c.fill=s["fh"]; c.alignment=s["ce"]; c.border=s["br"]
    ws.row_dimensions[1].height = 35
    col_widths = [22, 40, 32, 24, 50, 12, 20, 20, 20, 22, 45, 18]
    for ci, w in enumerate(col_widths[:len(HEADERS)], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def append_bids(ws, bids, s):
    from openpyxl.styles import PatternFill
    existing = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    added = 0
    for bid in bids:
        bn = str(bid["Bid Number"]).strip()
        if bn in existing:
            continue
        nr  = ws.max_row + 1
        org = bid.get("Organisation Name", "")
        fl  = s["fg"] if is_target_org(org) else (
              s["fa"] if nr % 2 == 0 else PatternFill("solid", fgColor=COL_WHT))
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(nr, ci, bid.get(h, "N/A"))
            c.font=s["df"]; c.fill=fl; c.border=s["br"]
            c.alignment = s["le"] if ci in (2, 3, 4, 5, 11) else s["ce"]
        existing.add(bn)
        added += 1
    return added


def rebuild_summary(wb, s, run_stats):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    ws["A1"] = "GeM Bids — Monthly Summary"
    ws["A1"].font = Font(bold=True, name="Arial", size=14, color=COL_HDR)
    ws["A2"] = f"Last updated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    filt = (f"Sector filter: {', '.join(TARGET_ORGANISATIONS[:5])}..."
            if TARGET_ORGANISATIONS else "Sector filter: ALL")
    ws["A3"] = filt
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells("A1:F1"); ws.merge_cells("A2:F2"); ws.merge_cells("A3:F3")

    for ci, h in enumerate(["Month","Total Bids","Sector Matched","Added This Run","Keywords Used","Sheet"], 1):
        c = ws.cell(5, ci, h)
        c.font=s["hf"]; c.fill=s["fs"]; c.alignment=s["ce"]; c.border=s["br"]

    sheets = [n for n in wb.sheetnames if n != "Summary"]
    grand  = 0
    for ri, name in enumerate(sheets, 6):
        wm    = wb[name]
        count = max(wm.max_row - 1, 0)
        kws   = {str(r[9]) for r in wm.iter_rows(min_row=2, values_only=True) if r[9]}
        sec   = sum(1 for r in wm.iter_rows(min_row=2, values_only=True)
                    if r[1] and is_target_org(str(r[1])))
        fl    = s["fa"] if ri % 2 == 0 else PatternFill("solid", fgColor=COL_WHT)
        for ci, v in enumerate([name, count, sec, run_stats.get(name, 0),
                                 ", ".join(sorted(kws)) or "—", f"-> '{name}'"], 1):
            c = ws.cell(ri, ci, v)
            c.font=s["df"]; c.fill=fl; c.border=s["br"]
            c.alignment = s["le"] if ci == 5 else s["ce"]
        grand += count

    tr = 6 + len(sheets)
    for ci, v in enumerate(["GRAND TOTAL", grand, "", "", "", ""], 1):
        c = ws.cell(tr, ci, v)
        c.font=s["bf"]; c.fill=s["ft"]; c.border=s["br"]; c.alignment=s["ce"]

    for ci, w in enumerate([18, 14, 16, 16, 50, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))


def save_excel(bids, output_path, run_stats):
    from openpyxl import Workbook, load_workbook
    month = datetime.now().strftime("%B %Y")
    s     = get_styles()

    if Path(output_path).exists():
        print(f"\n  Opening existing file : {output_path}")
        wb = load_workbook(output_path)
        if month in wb.sheetnames:
            wm = wb[month]
            old_headers = [wm.cell(1, c).value for c in range(1, wm.max_column + 1)]
            if "EMD Amount (₹)" in old_headers or wm.max_column != len(HEADERS):
                print(f"  Old schema in '{month}' — rebuilding...")
                del wb[month]
    else:
        print(f"\n  Creating new file     : {output_path}")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if month not in wb.sheetnames:
        ws = wb.create_sheet(month)
        write_headers(ws, s)
        print(f"  New sheet created     : '{month}'")
    else:
        ws = wb[month]
        print(f"  Sheet found           : '{month}'")

    added   = append_bids(ws, bids, s)
    skipped = len(bids) - added
    run_stats[month] = added
    print(f"  New bids added        : {added}")
    print(f"  Duplicates skipped    : {skipped}")
    rebuild_summary(wb, s, run_stats)

    # Auto-retry if Excel file is open
    max_retries = 10
    for attempt in range(1, max_retries + 1):
        try:
            wb.save(output_path)
            print(f"  Saved                 : {output_path}")
            break
        except PermissionError:
            if attempt == 1:
                print("")
                print(f"  ⚠️  '{Path(output_path).name}' is open in Excel.")
                print(f"  Please CLOSE it now. Retrying every 5 seconds...")
            msg = f"  Waiting... attempt {attempt}/{max_retries}  "
            print(msg, end="\r", flush=True)
            time.sleep(5)
            if attempt == max_retries:
                fallback = str(output_path).replace(
                    ".xlsx", f"_backup_{datetime.now().strftime('%H%M%S')}.xlsx"
                )
                wb.save(fallback)
                print("")
                print(f"  Could not save after {max_retries} attempts.")
                print(f"  Saved backup: {fallback}")
    return added


# ─── MAIN — SMART PIPELINE ─────────────────────────────────────────────────────

def main():
    now = datetime.now()
    print("=" * 62)
    print("  GeM Portal Bid Scraper  v7  —  Smart Filter-First Pipeline")
    print("=" * 62)
    print(f"  Output file  : {OUTPUT_FILE}")
    print(f"  Sheet        : '{now.strftime('%B %Y')}'")
    print(f"  PDF download : Only for bids passing card pre-filter + not in Excel")
    if TARGET_ORGANISATIONS:
        print(f"  Sector filter: {', '.join(TARGET_ORGANISATIONS[:3])}...")
    else:
        print(f"  Sector filter: ALL organisations")
    print("=" * 62)

    for pkg in ["selenium", "openpyxl", "webdriver_manager", "bs4", "lxml", "pypdf", "requests"]:
        try:
            __import__(pkg)
        except ImportError:
            print(f"\n  Missing: {pkg}")
            print("  Run: pip install selenium openpyxl webdriver-manager beautifulsoup4 lxml pypdf requests")
            sys.exit(1)

    # ── SCRAPING ────────────────────────────────────────────────────────────────
    all_bids, driver, kw_index = [], None, 0

    while kw_index < len(KEYWORDS):
        kw = KEYWORDS[kw_index]
        if driver is None:
            print("\n  Starting browser...")
            driver = make_driver()
        try:
            bids = search_keyword(driver, kw, MAX_PAGES_PER_KEYWORD)
            all_bids.extend(bids)
            if bids:
                print(f"  '{kw}': {len(bids)} bids")
            kw_index += 1
            time.sleep(1)
        except SessionDead:
            print(f"\n  Browser crashed. Restarting...")
            try: driver.quit()
            except Exception: pass
            driver = None
            time.sleep(3)
        except KeyboardInterrupt:
            print("\n  Interrupted. Processing collected data...")
            break
        except Exception as e:
            print(f"\n  Error on '{kw}': {e}")
            kw_index += 1

    if driver:
        try: driver.quit()
        except Exception: pass

    # ── STEP 1: Deduplicate within this run ─────────────────────────────────────
    seen, unique = set(), []
    for b in all_bids:
        if b["Bid Number"] not in seen:
            seen.add(b["Bid Number"]); unique.append(b)
    print(f"\n{'='*62}")
    print(f"  PIPELINE SUMMARY")
    print(f"{'='*62}")
    print(f"  Step 1 — Scraped & deduplicated : {len(unique)}")

    if not unique:
        print("  No bids found. Try HEADLESS=False or increase AJAX_WAIT.")
        return

    # ── STEP 2: Card pre-filter (no PDF download) ───────────────────────────────
    if TARGET_ORGANISATIONS:
        pre_filtered = [b for b in unique if card_prefilter(b)]
        dropped_by_card = len(unique) - len(pre_filtered)
        print(f"  Step 2 — Card pre-filter        : {len(pre_filtered)} kept / {dropped_by_card} dropped early (no PDF wasted)")
    else:
        pre_filtered = unique
        print(f"  Step 2 — Card pre-filter        : skipped (no sector filter active)")

    # ── STEP 3: Remove bids already in Excel ────────────────────────────────────
    existing_in_excel = get_existing_bid_numbers(OUTPUT_FILE)
    new_candidates    = [b for b in pre_filtered if b["Bid Number"] not in existing_in_excel]
    already_saved     = len(pre_filtered) - len(new_candidates)
    print(f"  Step 3 — Already in Excel       : {already_saved} skipped")
    print(f"  Step 3 — New candidates         : {len(new_candidates)}")

    if not new_candidates:
        print(f"\n  Nothing new to add. Excel is up to date.")
        return

    # ── STEP 4: Download PDFs only for new candidates ───────────────────────────
    if DOWNLOAD_BID_PDF:
        print(f"\n  Step 4 — PDF download           : {len(new_candidates)} bids  "
              f"({PARALLEL_PDF_WORKERS} parallel workers)")
        print(f"           (saved {len(unique) - len(new_candidates)} unnecessary downloads vs old method)")
        new_candidates = enrich_bids_parallel(new_candidates, max_workers=PARALLEL_PDF_WORKERS)

    # ── STEP 5: Final org-name filter using PDF Organisation Name ───────────────
    if TARGET_ORGANISATIONS:
        matched  = [b for b in new_candidates if is_target_org(b.get("Organisation Name", ""))]
        excluded = len(new_candidates) - len(matched)
        print(f"  Step 5 — Final sector filter    : {len(matched)} matched / {excluded} excluded")
        final_bids = matched if matched else new_candidates
        if not matched:
            print("  No sector matches — saving all candidates to avoid data loss.")
    else:
        final_bids = new_candidates
        print(f"  Step 5 — Sector filter          : skipped (saving all)")

    # ── STEP 5b: Block irrelevant Item Categories ────────────────────────────
    before_block = len(final_bids)
    final_bids   = [b for b in final_bids if not is_blocked_category(b.get("Item Category", ""))]
    blocked_count = before_block - len(final_bids)
    if blocked_count:
        print(f"  Step 5b— Blocked categories     : {blocked_count} removed (repair/taxi/cleaning etc.)")
    else:
        print(f"  Step 5b— Blocked categories     : none found")

    print(f"{'='*62}")

    # ── STEP 6: Save to Excel ───────────────────────────────────────────────────
    run_stats = {}
    added = save_excel(final_bids, OUTPUT_FILE, run_stats)

    print(f"\n{'='*62}")
    print(f"  Done!  {added} new bids added -> '{now.strftime('%B %Y')}' tab")
    print(f"  File : {OUTPUT_FILE}")
    print(f"  Green rows = sector-matched")
    print(f"{'='*62}")


if __name__ == "__main__":
    main()
